#!/usr/bin/env python3
# HTML 표 파싱과 Excel 저장에 공통으로 쓰는 유틸리티

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import TypeAlias

from bs4 import BeautifulSoup, NavigableString, Tag
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# HTML table 블록 탐색
TABLE_RE = re.compile(r"<table\b.*?</table>", re.IGNORECASE | re.DOTALL)

SHEET_TITLE_MAX_LENGTH = 31
SHEET_TITLE_INVALID_RE = re.compile(r"[\[\]\\/*?:]")

# 0-based Excel 병합 범위
MergeRange = tuple[int, int, int, int]
ParsedGrid = tuple[list[list[str]], list[MergeRange], list[bool]]
NumberValue: TypeAlias = int | float
CellValue: TypeAlias = str | NumberValue
NumericColumnFormats: TypeAlias = dict[int, str]

NUMERIC_RE = re.compile(
    r"^[+-]?(?:(?:\d+|\d{1,3}(?:,\d{3})+)(?:\.\d+)?|\.\d+)$"
)
MISSING_NUMERIC_TEXTS = {"-", "–", "—"}


@dataclass(frozen=True)
class ParsedCell:
    row: int
    col: int
    text: str


# 파일명 사용 불가 문자 정리
def sanitize_filename(text: str) -> str:
    text = text.replace("\t", " ")
    text = re.sub(r'[\/\\:*?"<>|]', "_", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:180].rstrip()


# 중복 파일명 suffix 처리
def unique_filename(base: str, used_names: dict[str, int]) -> str:
    used_names[base] = used_names.get(base, 0) + 1
    if used_names[base] == 1:
        return base
    return f"{base}_{used_names[base]}"


# Excel 시트명 제약에 맞춰 기존 표 이름을 보정
def sanitize_sheet_title(text: str) -> str:
    title = SHEET_TITLE_INVALID_RE.sub("_", text).strip("'").strip()
    return title or "Sheet"


# 한 워크북 안에서 중복되지 않는 시트명 생성
def unique_sheet_title(base: str, used_titles: set[str]) -> str:
    base = sanitize_sheet_title(base)

    for suffix_index in range(1, len(used_titles) + 2):
        suffix = "" if suffix_index == 1 else f"_{suffix_index}"
        title = f"{base[:SHEET_TITLE_MAX_LENGTH - len(suffix)]}{suffix}"
        normalized_title = title.casefold()
        if normalized_title not in used_titles:
            used_titles.add(normalized_title)
            return title

    raise RuntimeError("시트명을 만들 수 없습니다.")


# HTML 셀 텍스트 추출
def get_cell_text(cell: Tag) -> str:
    parts: list[str] = []
    for element in cell.descendants:
        if isinstance(element, NavigableString):
            parts.append(str(element))
        elif isinstance(element, Tag) and element.name == "br":
            parts.append("\n")

    lines = [
        re.sub(r"[ \t]+", " ", line).strip()
        for line in "".join(parts).split("\n")
    ]
    return "\n".join(lines).strip()


# rowspan/colspan 안전 변환
def _get_span(cell: Tag, attr: str) -> int:
    try:
        return max(1, int(cell.get(attr, 1)))
    except (TypeError, ValueError):
        return 1


# rowspan 점유 열 건너뛰기
def _next_available_column(row: int, start_col: int, occupied: set[tuple[int, int]]) -> int:
    col = start_col
    while (row, col) in occupied:
        col += 1
    return col


# 병합 셀 점유 좌표 표시
def _mark_spanned_cells(
    occupied: set[tuple[int, int]],
    start_row: int,
    start_col: int,
    rowspan: int,
    colspan: int,
) -> None:
    for row in range(start_row, start_row + rowspan):
        for col in range(start_col, start_col + colspan):
            if (row, col) != (start_row, start_col):
                occupied.add((row, col))


# HTML table을 Excel용 그리드로 변환
def parse_table_to_grid(table_html: str) -> ParsedGrid:
    soup = BeautifulSoup(table_html, "html.parser")
    table = soup.find("table")
    if table is None:
        return [], [], []

    rows = table.find_all("tr")
    placed_cells: list[ParsedCell] = []
    occupied: set[tuple[int, int]] = set()
    merges: list[MergeRange] = []
    is_header = [False] * len(rows)
    max_rows = len(rows)
    max_cols = 0

    for row_index, tr in enumerate(rows):
        col_index = 0
        for cell in tr.find_all(["th", "td"]):
            if cell.name == "th":
                is_header[row_index] = True

            col_index = _next_available_column(row_index, col_index, occupied)
            rowspan = _get_span(cell, "rowspan")
            colspan = _get_span(cell, "colspan")

            placed_cells.append(
                ParsedCell(
                    row=row_index,
                    col=col_index,
                    text=get_cell_text(cell),
                )
            )
            _mark_spanned_cells(occupied, row_index, col_index, rowspan, colspan)

            if rowspan > 1 or colspan > 1:
                merges.append(
                    (
                        row_index,
                        col_index,
                        row_index + rowspan - 1,
                        col_index + colspan - 1,
                    )
                )

            max_rows = max(max_rows, row_index + rowspan)
            max_cols = max(max_cols, col_index + colspan)
            col_index += colspan

    is_header.extend([False] * (max_rows - len(is_header)))

    grid = [[""] * max_cols for _ in range(max_rows)]
    for cell in placed_cells:
        grid[cell.row][cell.col] = cell.text

    return grid, merges, is_header


# 숫자로 변환 가능한 셀 텍스트 파싱
def _parse_numeric_value(value: str) -> NumberValue | None:
    text = value.strip()
    if not text or not NUMERIC_RE.fullmatch(text):
        return None

    normalized = text.replace(",", "")
    try:
        return int(normalized)
    except ValueError:
        return float(normalized)


# 열마다 처음 나오는 숫자 셀부터 아래쪽 값을 숫자 값으로 변환
def convert_numeric_columns(
    grid: list[list[str]],
    is_header: list[bool],
) -> tuple[list[list[CellValue]], NumericColumnFormats]:
    if not grid:
        return [], {}

    converted: list[list[CellValue]] = [row[:] for row in grid]
    numeric_column_formats: NumericColumnFormats = {}
    max_cols = max((len(row) for row in grid), default=0)

    for col_index in range(max_cols):
        numeric_started = False
        has_float = False

        for row_index, row in enumerate(grid):
            if col_index >= len(row):
                continue

            value = row[col_index]
            if not value.strip() or value.strip() in MISSING_NUMERIC_TEXTS:
                continue

            number = _parse_numeric_value(value)
            if number is None:
                continue

            numeric_started = True
            converted[row_index][col_index] = number
            if isinstance(number, float):
                has_float = True

        if numeric_started:
            numeric_column_formats[col_index] = "#,##0.00" if has_float else "#,##0"

    return converted, numeric_column_formats


# 셀 값과 기본 스타일 적용
def _apply_cell_values_and_styles(
    worksheet,
    grid: list[list[CellValue]],
    is_header: list[bool],
    numeric_column_formats: NumericColumnFormats,
) -> None:
    bold = Font(bold=True)
    align = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for row_index, row in enumerate(grid, start=1):
        for col_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = align
            if is_header[row_index - 1]:
                cell.font = bold
            elif isinstance(value, (int, float)):
                number_format = numeric_column_formats.get(col_index - 1)
                if number_format is not None:
                    cell.number_format = number_format


# 병합 범위 적용
def _apply_merges(worksheet, merges: list[MergeRange]) -> None:
    for start_row, start_col, end_row, end_col in merges:
        if start_row == end_row and start_col == end_col:
            continue
        worksheet.merge_cells(
            start_row=start_row + 1,
            start_column=start_col + 1,
            end_row=end_row + 1,
            end_column=end_col + 1,
        )


# 열 너비 자동 조정
def _autosize_columns(worksheet, grid: list[list[CellValue]]) -> None:
    if not grid:
        return

    for col_index in range(len(grid[0])):
        longest_line = max(
            (
                len(line)
                for row in grid
                if col_index < len(row)
                for line in str(row[col_index]).split("\n")
            ),
            default=8,
        )
        worksheet.column_dimensions[get_column_letter(col_index + 1)].width = min(
            max(longest_line + 2, 10),
            50,
        )


# 워크시트에 표 쓰기
def write_table_to_worksheet(
    workbook: Workbook,
    sheet_name: str,
    grid: list[list[str]],
    merges: list[MergeRange],
    is_header: list[bool],
) -> None:
    worksheet = workbook.create_sheet(title=sheet_name)
    converted_grid, numeric_column_formats = convert_numeric_columns(grid, is_header)

    _apply_cell_values_and_styles(
        worksheet,
        converted_grid,
        is_header,
        numeric_column_formats,
    )
    _apply_merges(worksheet, merges)
    _autosize_columns(worksheet, converted_grid)


# 단일 표 Excel 파일 저장
def write_excel(
    grid: list[list[str]],
    merges: list[MergeRange],
    is_header: list[bool],
    output_path: Path,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    write_table_to_worksheet(workbook, "Sheet1", grid, merges, is_header)

    workbook.save(output_path)
