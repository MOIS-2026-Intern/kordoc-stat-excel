#!/usr/bin/env python3
# 통계연보 마크다운의 HTML 표를 Excel 파일로 저장
# 공용 표 파싱/Excel 저장 함수는 범용 추출기에서도 재사용

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from bs4 import BeautifulSoup, NavigableString, Tag
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

INPUT_FILE = Path("/Users/song/dev/mois/kordoc/통계연보파싱.md")
OUTPUT_DIR = Path("/Users/song/dev/mois/kordoc/statoutput")

# 통계연보 코드형 헤딩 탐색
HEADING_RE = re.compile(r"^(\d+(?:-\d+)+)\t(.+)$", re.MULTILINE)

# HTML table 블록 탐색
TABLE_RE = re.compile(r"<table\b.*?</table>", re.IGNORECASE | re.DOTALL)

# 0-based Excel 병합 범위
MergeRange = tuple[int, int, int, int]
ParsedGrid = tuple[list[list[str]], list[MergeRange], list[bool]]


@dataclass(frozen=True)
class TableMatch:
    heading: str | None
    html: str
    line_no: int


@dataclass(frozen=True)
class ParsedCell:
    row: int
    col: int
    text: str


# 파일명 사용 불가 문자 정리
def sanitize_filename(text: str) -> str:
    text = text.replace("\t", " ")
    text = re.sub(r'[\/\\:*?"<>|]', "_", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:180].rstrip()


# HTML 셀 텍스트 추출
def get_cell_text(cell: Tag) -> str:
    parts: list[str] = []
    for element in cell.descendants:
        if isinstance(element, NavigableString):
            parts.append(str(element))
        elif isinstance(element, Tag) and element.name == "br":
            parts.append("\n")

    lines = [
        re.sub(r"[ \t]+", " ", line).strip()
        for line in "".join(parts).split("\n")
    ]
    return "\n".join(lines).strip()


# rowspan/colspan 안전 변환
def _get_span(cell: Tag, attr: str) -> int:
    try:
        return max(1, int(cell.get(attr, 1)))
    except (TypeError, ValueError):
        return 1


# rowspan 점유 열 건너뛰기
def _next_available_column(row: int, start_col: int, occupied: set[tuple[int, int]]) -> int:
    col = start_col
    while (row, col) in occupied:
        col += 1
    return col


# 병합 셀 점유 좌표 표시
def _mark_spanned_cells(
    occupied: set[tuple[int, int]],
    start_row: int,
    start_col: int,
    rowspan: int,
    colspan: int,
) -> None:
    for row in range(start_row, start_row + rowspan):
        for col in range(start_col, start_col + colspan):
            if (row, col) != (start_row, start_col):
                occupied.add((row, col))


# HTML table을 Excel용 그리드로 변환
def parse_table_to_grid(table_html: str) -> ParsedGrid:
    soup = BeautifulSoup(table_html, "html.parser")
    table = soup.find("table")
    if table is None:
        return [], [], []

    rows = table.find_all("tr")
    placed_cells: list[ParsedCell] = []
    occupied: set[tuple[int, int]] = set()
    merges: list[MergeRange] = []
    is_header = [False] * len(rows)
    max_rows = len(rows)
    max_cols = 0

    for row_index, tr in enumerate(rows):
        col_index = 0
        for cell in tr.find_all(["th", "td"]):
            if cell.name == "th":
                is_header[row_index] = True

            col_index = _next_available_column(row_index, col_index, occupied)
            rowspan = _get_span(cell, "rowspan")
            colspan = _get_span(cell, "colspan")

            placed_cells.append(
                ParsedCell(
                    row=row_index,
                    col=col_index,
                    text=get_cell_text(cell),
                )
            )
            _mark_spanned_cells(occupied, row_index, col_index, rowspan, colspan)

            if rowspan > 1 or colspan > 1:
                merges.append(
                    (
                        row_index,
                        col_index,
                        row_index + rowspan - 1,
                        col_index + colspan - 1,
                    )
                )

            max_rows = max(max_rows, row_index + rowspan)
            max_cols = max(max_cols, col_index + colspan)
            col_index += colspan

    is_header.extend([False] * (max_rows - len(is_header)))

    grid = [[""] * max_cols for _ in range(max_rows)]
    for cell in placed_cells:
        grid[cell.row][cell.col] = cell.text

    return grid, merges, is_header


# 셀 값과 기본 스타일 적용
def _apply_cell_values_and_styles(
    worksheet,
    grid: list[list[str]],
    is_header: list[bool],
) -> None:
    bold = Font(bold=True)
    align = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for row_index, row in enumerate(grid, start=1):
        for col_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = align
            if is_header[row_index - 1]:
                cell.font = bold


# 병합 범위 적용
def _apply_merges(worksheet, merges: list[MergeRange]) -> None:
    for start_row, start_col, end_row, end_col in merges:
        if start_row == end_row and start_col == end_col:
            continue
        worksheet.merge_cells(
            start_row=start_row + 1,
            start_column=start_col + 1,
            end_row=end_row + 1,
            end_column=end_col + 1,
        )


# 열 너비 자동 조정
def _autosize_columns(worksheet, grid: list[list[str]]) -> None:
    if not grid:
        return

    for col_index in range(len(grid[0])):
        longest_line = max(
            (
                len(line)
                for row in grid
                if col_index < len(row)
                for line in row[col_index].split("\n")
            ),
            default=8,
        )
        worksheet.column_dimensions[get_column_letter(col_index + 1)].width = min(
            max(longest_line + 2, 10),
            50,
        )


# Excel 파일 저장
def write_excel(
    grid: list[list[str]],
    merges: list[MergeRange],
    is_header: list[bool],
    output_path: Path,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    _apply_cell_values_and_styles(worksheet, grid, is_header)
    _apply_merges(worksheet, merges)
    _autosize_columns(worksheet, grid)

    workbook.save(output_path)


# 표와 가장 가까운 앞쪽 헤딩 매칭
def find_tables_with_headings(text: str) -> list[TableMatch]:
    headings = [(match.start(), match.group(0)) for match in HEADING_RE.finditer(text)]
    matches: list[TableMatch] = []

    for table_match in TABLE_RE.finditer(text):
        heading = next(
            (
                heading_text
                for heading_start, heading_text in reversed(headings)
                if heading_start < table_match.start()
            ),
            None,
        )
        line_no = text.count("\n", 0, table_match.start()) + 1
        matches.append(
            TableMatch(
                heading=heading,
                html=table_match.group(0),
                line_no=line_no,
            )
        )

    return matches


# 중복 파일명 suffix 처리
def _unique_filename(base: str, used_names: dict[str, int]) -> str:
    used_names[base] = used_names.get(base, 0) + 1
    if used_names[base] == 1:
        return base
    return f"{base}_{used_names[base]}"


# 통계연보 마크다운의 표를 Excel로 변환
def convert_md_to_excel(md_path: Path, output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    text = md_path.read_text(encoding="utf-8")

    used_names: dict[str, int] = {}
    saved_paths: list[Path] = []

    for table in find_tables_with_headings(text):
        if table.heading is None:
            continue

        base_name = sanitize_filename(table.heading)
        file_stem = _unique_filename(base_name, used_names)
        output_path = output_dir / f"{file_stem}.xlsx"

        try:
            grid, merges, is_header = parse_table_to_grid(table.html)
            if not grid:
                continue

            write_excel(grid, merges, is_header, output_path)
            saved_paths.append(output_path)
        except Exception as error:
            print(f"  [error] line {table.line_no} ({file_stem}): {error}")

    return saved_paths


# CLI 실행 진입점
def main() -> None:
    saved = convert_md_to_excel(INPUT_FILE, OUTPUT_DIR)
    print(f"저장: {len(saved)}개")
    print(f"출력 폴더: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
